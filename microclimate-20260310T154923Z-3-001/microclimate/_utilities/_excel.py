# -*- coding: utf-8 -*-
"""
Created on Fri Dec  9 10:45:09 2022

Module with functions to interact with excel spreadsheets

@author: Viet.Le
"""

import logging
import openpyxl
import numpy as np
import pandas as pd
from ._load import validFile

logger = logging.getLogger(__name__)



#%% writing to excel

def writeEXCEL(datadf,filename,sheetname,appendmode='w'):
    """
    Export QA information in dataframe to excel and color

    Parameters
    ----------
    datadf : dataframe
        dataframe containing the percentile winds and contributions.
    filename : TYPE
        DESCRIPTION.
    sheetname : TYPE
        DESCRIPTION.
    appendmode : TYPE, optional
        DESCRIPTION. The default is 'w'.

    Returns
    -------
    None.

    """

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode=appendmode)
    datadf.to_excel(writer, sheet_name=sheetname, index=False)

    # cell formats
    ws = writer.sheets[sheetname]
    ws.freeze_panes = "A2"

    for col in ws.iter_cols():
        if col[0].value in ['x','y','z'] or isinstance(col[0].value, int):
            continue
        elif col[0].value in ['Rank']:
            for cell in col:
                if cell.value == 1:
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = '08fc04')
                elif cell.value == 2:
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'FFFC04')
                elif cell.value == 3:
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'ff8404')
                elif cell.value == 4:
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'ff0404')
                elif cell.value == 5:
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'ff04bc')
        elif col[0].value in ['U50','U95','U99p975','Total - 95']:

            if col[0].value == 'U95':
                for cell in col[1::]:
                    cell.number_format = '0.00'
                    if cell.value <= 4:
                        cell.fill = openpyxl.styles.PatternFill('solid', fgColor = '00AA00')
                    elif cell.value > 4 and cell.value <= 6:
                        cell.fill = openpyxl.styles.PatternFill('solid', fgColor = '91D14F')
                    elif cell.value > 6 and cell.value <= 8:
                        cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'FFDE52')
                    elif cell.value > 8 and cell.value <= 10:
                        cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'FF9900')
                    elif cell.value > 10:
                        cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'FF0000')
            elif col[0].value == 'U99p975':
                for cell in col[1::]:
                    cell.number_format = '0.00'
                    if cell.value >= 15:
                        cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'FF0000')

            elif col[0].value in ['Total - 95']:
                for cell in col:
                    cell.number_format = '0.00'
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'e0e4f4')

            else:
                for cell in col:
                    cell.number_format = '0.00'

        elif col[0].value in ['Total 99.975']:
            for cell in col:
                cell.number_format = '0.0000'
                cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'e0e4f4')

        elif col[0].value[-4::] in ['comf']:
            for cell in col[1::]:
                if cell.value > 1:
                    cell.fill = openpyxl.styles.PatternFill('solid', fgColor = 'fcc4d0')
                    # cell.font = openpyxl.styles.Font(color = 'ff5c54')

    # writer.save()
    writer.close()



#%% Functions for interacting with ESDU spreadsheet

def load_and_prep_ESDUwindspeeds(path2ESDU):
    """
    loads in an ESDU spreadsheet and extracts only the mean and gust wind 
    speeds

    Parameters
    ----------
    path2ESDU : str
        path to ESDU.

    Raises
    ------
    Exception
        Error raised when there are more columns in the ESDU than standard wind 
        directions.

    Returns
    -------
    MWS : dataframe
        mean wind speeds from ESDU.
    GWS : dataframe
        gust wind speeds from ESDU.

    """
    
    if not validFile(path2ESDU):
        raise OSError(f"Path to {path2ESDU} not found")
    
    # load in the excel spreadsheet with the ESDU wind speeds
    pd_excel = pd.read_excel(path2ESDU,sheet_name='Wind speed',header=7)

    # clean the ESDU spreadsheet to extract the needed data
    dfclean = clean_ESDU_sprdsht(pd_excel)
    U = dfclean

    # split U data into two: one for mean winds and the other for gust winds
    num_heights = int(len(U))
    MWS = U.iloc[int(0):int(num_heights/2)].reset_index(drop=True)
    GWS = U.iloc[int(num_heights/2)::].reset_index(drop=True)

    # rename the columns to the wind directions
    num_columns = len(U.columns)
    column_indices = range(0,num_columns,1)

    # check number of columns in the ESDU
    if num_columns == 13:
        new_names  = range(0,360,30)
    elif num_columns == 17:
        new_names  = np.arange(0,360,22.5)
    elif num_columns == 9:
        new_names  = range(0,360,45)
    else:
        logger.error(f'There were more columns than standard wind directions in the ESDU: {path2ESDU}')
        raise Exception(f'You do not have exactly 8, 12, or 16 wind directions in this ESDU!\n{path2ESDU}')

    old_names = U.columns[column_indices][1::]
    dict_colnames = dict(zip(old_names, new_names))
    dict_colnames['Height above ground (m)'] = 'z'
    MWS.rename(columns=dict_colnames, inplace=True)
    GWS.rename(columns=dict_colnames, inplace=True)

    return MWS, GWS



def load_and_prep_ESDUturbprop(path2ESDU,sheetname_turb='u-component turbulence'):
    """
    loads in an ESDU spreadsheet and extracts only the turbulence intensity and
    the length scales in x, y, z directions

    Parameters
    ----------
    path2ESDU : str
        path to ESDU.
    sheetname_turb : str, optional
        name of the sheet in the ESDU that holds the turbulence information.

    Raises
    ------
    OSError
        DESCRIPTION.
    Exception
        DESCRIPTION.

    Returns
    -------
    I : dataframe
        turbulence intensity.
    Lx : dataframe
        turbulence length scale in x-direction.
    Ly : dataframe
        turbulence length scale in y-direction.
    Lz : dataframe
        turbulence length scale in z-direction.

    """
    
    if not validFile(path2ESDU):
        raise OSError(f"Path to {path2ESDU} not found")
        
    # load in the excel spreadsheet with the ESDU wind speeds
    pd_excel = pd.read_excel(path2ESDU,sheet_name=sheetname_turb,header=7)

    # clean the ESDU spreadsheet to extract the needed data
    dfclean = clean_ESDU_sprdsht(pd_excel)

    # rename the columns to the wind directions
    num_columns = len(dfclean.columns)
    column_indices = range(0,num_columns,1)

    # check number of columns in the ESDU
    if num_columns == 13:
        new_names  = range(0,360,30)
    elif num_columns == 17:
        new_names  = np.arange(0,360,22.5)
    elif num_columns == 9:
        new_names  = range(0,360,45)
    else:
        logger.error(f'There were more columns than standard wind directions in the ESDU: {path2ESDU}')
        raise Exception(f'You do not have exactly 8, 12, or 16 wind directions in this ESDU!\n{path2ESDU}')

    # split U data into four: I, Lx, Ly, Lz for the specified wind-direction
    dfsplit = np.array_split(dfclean,4)
    I = dfsplit[0]
    Lx = dfsplit[1]
    Ly = dfsplit[2]
    Lz = dfsplit[3]

    old_names = I.columns[column_indices][1::]
    dict_colnames = dict(zip(old_names, new_names))
    dict_colnames['Height above ground (m)'] = 'z'
    I.rename(columns=dict_colnames, inplace=True)
    Lx.rename(columns=dict_colnames, inplace=True)
    Ly.rename(columns=dict_colnames, inplace=True)
    Lz.rename(columns=dict_colnames, inplace=True)

    return I, Lx, Ly, Lz



def clean_ESDU_sprdsht(pd_excel):
    """
    

    Parameters
    ----------
    pd_excel : dataframe
        raw dataframe from loading in the ESDU information directly.

    Returns
    -------
    df : dataframe
        cleaned ESDU dataframe.

    """

    # find out which cell entries are strings in the column named 'Height above ground (m)' by default
    
    def check_nan_or_string(value):
        if isinstance(value, str):
            return True
        elif isinstance(value, float) and np.isnan(value):
            return True
        else:
            return False
    idx_strings = pd_excel['Height above ground (m)'].apply(lambda x: check_nan_or_string(x))

    # extract only the numeric data
    df = pd_excel[~idx_strings]

    # convert everything to flaot
    df = df.astype(float)

    # drop any rows with all NaNs, then repeat with columns
    df = df.dropna(axis=0,how='all')
    df = df.dropna(axis=1,how='all')

    return df



#%% Transposition factor

def calcTF(Uap,Usite,refheight=30,refWS=10):

    """
    Calculates the transposition factors

    Parameters
    ----------
    Uap :
        Dataframe of ESDU airport wind speeds and elevations
    Usite :
        Dataframe of ESDU site wind speeds and elevations
    refheight :
        reference height to calculate the TF. Default is 30 m
    refWS :
        reference wind speed. Default is 10 m/s

    Returns
    -------
    TF_refheight :
        Combined transposition factors at ref height
    ap_OC_at_10m :
        Transposition factors from airport to OC at 10m
    OC_site_at_refheight :
        Transposition factors from OC to site at selected reference height

    """

    # transition factors all heights
    TF_ap_to_OC = Uap.iloc[:,1::] / refWS
    TF_OC_to_site = Usite.iloc[:,1::] / refWS

    # transition factor at specific heights
    ap_OC_at_10m = TF_ap_to_OC[Uap['z'] == 10]
    OC_site_at_refheight = TF_OC_to_site[Uap['z'] == refheight]

    # get the combined transition factors at the reference height
    TF_refheight = ap_OC_at_10m.to_numpy() * OC_site_at_refheight.to_numpy()
    
    logger.info("Transposition factors calculated")

    return TF_refheight, ap_OC_at_10m, OC_site_at_refheight

    
