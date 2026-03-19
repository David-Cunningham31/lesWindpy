# -*- coding: utf-8 -*-
"""
Created on Fri Dec  9 11:29:20 2022

@author: Viet.Le
"""

import logging
import numpy as np
import pandas as pd
from tqdm import tqdm

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


#%%

def importWeibulls(pathWeibull,season='annual',num_dir=12):
    if season == 'annual':
        numheader = 6
    elif season == 'autumn' or season == 'fall':
        numheader = 10
    elif season == 'winter':
        numheader = 14
    elif season == 'spring':
        numheader = 18
    elif season == 'summer':
        numheader = 22
    

    if num_dir == 12:
        cols_to_use = "B:M"
    elif num_dir == 16:
        cols_to_use = "B:Q"
    elif num_dir == 8:
        cols_to_use = "B:I"
    else:
        raise Exception("Please enter 8, 12, or 16 for number of directions")
        
    Weibulls = pd.read_excel(pathWeibull,
                             sheet_name='Sheet1',
                             header=numheader,
                             usecols=cols_to_use,
                             nrows=3).transpose()
    Weibulls = Weibulls.rename(columns={0: "p", 1: "c", 2: "k"})

    p = Weibulls['p'].to_numpy() # freq
    c = Weibulls['c'].to_numpy() # scale, aka lambda
    k = Weibulls['k'].to_numpy() # shape
    logger.debug(f'{season} weibulls have been used\n')

    return p, c, k



def VC_filter(data):
    """
    Visual Crossing filter to keep only a single data entry per provided hour.
    Details of the filter are provided here:
    https://www.visualcrossing.com/resources/documentation/weather-data/how-we-process-integrated-surface-database-historical-weather-data/

    Parameters
    ----------
    data : pandas dataframe
        data to be filtered. Please ensure outliers have already been removed

    Returns
    -------
    data_cleaned_VC : pandas dataframe
        filtered dataframe using VC method. Only one data entry per provided hour

    """
    
    
    try:
        if (data['WindSpeed'] >= 100).any():
            raise Exception("Please remove all outlier wind speeds (greater than 100 m/s) before applying the VC filter!")
        elif (data['WindDirection'] >= 360).any():
            raise Exception("Please remove all outlier wind directions (greater than 360 degrees) before applying the VC filter!")
    except KeyError:
        if (data['WindDir'] >= 360).any():
            raise Exception("Please remove all outlier wind directions (greater than 360 degrees) before applying the VC filter!")
        
        
    """
    Move information beyond the 30 minute mark to the next hour
    """
    def adjust_hour_past30min(x):
        if (30 < x.minute < 60):
            x += pd.DateOffset(minutes=60-(x.minute % 30)*2)
        return x
    data.index = data.index.map(adjust_hour_past30min)



    """
    Rank data based on how far away it is from the hour
    """
    def rank_closest_to_hour(x):
        minute = x.minute
        if (0 <= minute <= 10) or (50 <= minute <= 59):
            rank_quality = 1
        elif (10 < minute <= 20) or (40 < minute < 50):
            rank_quality = 0.5
        elif (20 < minute <= 40):
            rank_quality = 0
        else:
            rank_quality = 0
        return rank_quality
    data['RK_HR'] = data.index.map(rank_closest_to_hour)



    """
    Rank data based on Wind Speed Quality
    """
    def rank_quality(x):
        if x not in [2, 3, 6, 7]:
            rank_quality = 1
        else:
            rank_quality = 0
        return rank_quality
    data['RK_QC'] = data['QCWindSpeed'].apply(rank_quality)



    """
    Rank data based on Wind Speed Quality
    """
    def rank_quality(x):
        if x not in [2, 3, 6, 7]:
            rank_quality = 1
        else:
            rank_quality = 0
        return rank_quality
    data['RK_QC'] = data['QCWindSpeed'].apply(rank_quality)



    """
    Rank data based on Report Type Quality (merger of reports is better)
    """
    def rank_reporttype(x):
        if x in ['S-S-A','SA-AU','SY-AE','SY-AU','SY-MT']:
            rank_quality = 1
        else:
            rank_quality = 0
        return rank_quality
    data['RK_RT'] = data['ReportType'].apply(rank_reporttype)



    """
    Add all the ranks together
    """
    data['Rank'] = data['RK_HR'] + data['RK_QC'] + data['RK_RT']



    """
    select the hour that has the highest ranking
    """

    # make a new column that only has the hour to see which hours are repeated
    data['Hour'] = data.index.strftime('%Y-%m-%d-%H')
    def choose_rank(x):
        idx = np.argmax(x['Rank'])
        return x.iloc[idx]

    # for each group of duplicated hours, apply the function "choose rank"
    tqdm.pandas()
    data_cleaned_VC = data.groupby(by=["Hour"]).progress_apply(choose_rank)

    # convert the index to datetime and rename it to what it was originally
    data_cleaned_VC.index = pd.to_datetime(data_cleaned_VC.index)
    data_cleaned_VC.index.name = 'datetimelist'

    return data_cleaned_VC



def remove_varDir(data, remove=True):
    mask = data['WindDir'] > 360

    return _remove_masked(data, mask, remove)



def remove_threshold(data, threshold=100, remove=True):
    mask = data['WindSpeed'] > threshold

    return _remove_masked(data, mask, remove)



def convert_360(data,varname='WindDir'):
    mask = data[varname] == 360
    data_converted = data.copy(deep=True)
    data_converted.loc[mask,varname] = 0
    return data_converted, data[mask]



def _remove_masked(data, mask, remove):
    '''
    Removes masked data

    Parameters
    ----------
    data : pandas.DataFrame
        input pandas DataFrame.
    mask : pandas.Series
        Series of boolean values marking the samples to remove. True values
        are removed.
    remove : bool
        If True, the marked values are removed from the dataset inplace.
        Otherwise the values are returned, but are mainteined in the dataset.

    Returns
    -------
    pandas.DataFrame
        A dateset containing the marked samples

    '''

    # Store the samples to be returned
    out = data[mask]

    if remove:
        if sum(mask) > 0:
            # Drop the marked samples
            data_cleaned = data.drop(data[mask].index)

    return data_cleaned, out




#%% export functions

def writeEXCEL_wbparam(datadf,filename,appendmode='w',num_headerlines_skip=0):
    """
    
    """
    
    if appendmode=='w':
        with pd.ExcelWriter(filename, engine='openpyxl', mode=appendmode) as writer:
            datadf.to_excel(writer, sheet_name="Sheet1", startrow=num_headerlines_skip, 
                            index=True, header=True)
            
    else:
        workbook = load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl', mode=appendmode, if_sheet_exists='overlay') as writer:
            
            """
            older versions of pandas are unable to append to an existing worksheet.
            You'll need to add the follow line to append to existing
            """
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            
            # append the dataframe to the existing worksheet
            datadf.to_excel(writer, sheet_name="Sheet1", startrow=workbook['Sheet1'].max_row, 
                            index=True, header=True)



def writeEXCEL_exceedance(datadf,filename,num_headerlines_skip=0,addrow=0):
    """
    
    """
    workbook = load_workbook(filename)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        
        """
        older versions of pandas are unable to append to an existing worksheet.
        You'll need to add the follow line to append to existing
        """
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        
        # append the dataframe to the existing worksheet
        datadf.to_excel(writer, sheet_name="Sheet1", startrow=workbook['Sheet1'].max_row+addrow, 
                        index=True, header=True)




